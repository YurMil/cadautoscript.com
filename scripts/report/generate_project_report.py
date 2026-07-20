#!/usr/bin/env python3
"""Generate an Excel progress report (task table + Gantt chart) from a GitHub Project v2 board.

Fetches all items of the configured user project via the GitHub GraphQL API,
computes percent-complete from issue state and acceptance-criteria checkboxes,
and renders an .xlsx with a weekly Gantt timeline.

Environment variables:
    GITHUB_TOKEN   token with `read:project` scope (classic PAT: project + repo)
    PROJECT_OWNER  user login owning the project (default: YurMil)
    PROJECT_NUMBER project number (default: 2)
    OUTPUT_FILE    output path (default: project-report.xlsx)
"""

from __future__ import annotations

import datetime as dt
import json
import os
import re
import sys
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

API = "https://api.github.com/graphql"

QUERY = """
query($login: String!, $number: Int!, $after: String) {
  user(login: $login) {
    projectV2(number: $number) {
      title
      items(first: 100, after: $after) {
        pageInfo { hasNextPage endCursor }
        nodes {
          content {
            ... on Issue { number title state body url createdAt closedAt }
            ... on DraftIssue { title body }
          }
          fieldValues(first: 20) {
            nodes {
              ... on ProjectV2ItemFieldSingleSelectValue {
                name
                field { ... on ProjectV2SingleSelectField { name } }
              }
              ... on ProjectV2ItemFieldDateValue {
                date
                field { ... on ProjectV2Field { name } }
              }
            }
          }
        }
      }
    }
  }
}
"""

STATUS_FALLBACK_PCT = {
    "Done": 100,
    "In review": 90,
    "In progress": 50,
    "Ready": 10,
    "Backlog": 0,
    "To triage": 0,
}

STATUS_COLOR = {
    "Done": "4CAF50",
    "In review": "8BC34A",
    "In progress": "2196F3",
    "Ready": "FFC107",
    "Backlog": "B0BEC5",
    "To triage": "CFD8DC",
}


def gql(token: str, variables: dict) -> dict:
    body = json.dumps({"query": QUERY, "variables": variables}).encode()
    req = urllib.request.Request(
        API,
        data=body,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
    )
    with urllib.request.urlopen(req) as resp:
        data = json.load(resp)
    if "errors" in data:
        raise RuntimeError(f"GraphQL errors: {data['errors']}")
    return data["data"]


def fetch_items(token: str, login: str, number: int) -> tuple[str, list[dict]]:
    items, after = [], None
    title = ""
    while True:
        data = gql(token, {"login": login, "number": number, "after": after})
        project = data["user"]["projectV2"]
        title = project["title"]
        page = project["items"]
        items.extend(page["nodes"])
        if not page["pageInfo"]["hasNextPage"]:
            return title, items
        after = page["pageInfo"]["endCursor"]


def checkbox_pct(body: str | None) -> int | None:
    if not body:
        return None
    checked = len(re.findall(r"^\s*[-*] \[[xX]\]", body, re.M))
    total = checked + len(re.findall(r"^\s*[-*] \[ \]", body, re.M))
    return round(100 * checked / total) if total else None


def parse_item(node: dict) -> dict | None:
    content = node.get("content") or {}
    if not content.get("title"):
        return None
    fields: dict[str, str] = {}
    for fv in node["fieldValues"]["nodes"]:
        field = fv.get("field")
        if not field:
            continue
        fields[field["name"]] = fv.get("name") or fv.get("date")

    status = fields.get("Status", "To triage")
    state = content.get("state")
    if state == "CLOSED":
        pct = 100
    else:
        pct = checkbox_pct(content.get("body"))
        if pct is None or pct == 0:
            pct = STATUS_FALLBACK_PCT.get(status, 0)

    def to_date(value: str | None) -> dt.date | None:
        return dt.date.fromisoformat(value[:10]) if value else None

    start = to_date(fields.get("Start date")) or to_date(content.get("createdAt"))
    finish = to_date(fields.get("Target date")) or to_date(content.get("closedAt"))

    return {
        "number": content.get("number"),
        "title": content["title"],
        "url": content.get("url", ""),
        "status": status,
        "priority": fields.get("Priority", ""),
        "size": fields.get("Size", ""),
        "area": fields.get("Area", ""),
        "pct": pct,
        "start": start,
        "finish": finish,
    }


def build_workbook(project_title: str, rows: list[dict], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress Report"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="263238")
    head_font = Font(bold=True, color="FFFFFF")

    headers = ["ID", "Task Name", "Area", "Priority", "Size", "Status",
               "Start", "Finish", "% Complete"]
    n_static = len(headers)

    # Weekly timeline bounds
    dates = [r["start"] for r in rows if r["start"]] + [r["finish"] for r in rows if r["finish"]]
    today = dt.date.today()
    lo = min(dates) if dates else today
    hi = max(dates) if dates else today
    hi = max(hi, today)
    lo -= dt.timedelta(days=lo.weekday())  # align to Monday
    weeks = []
    week = lo
    while week <= hi:
        weeks.append(week)
        week += dt.timedelta(days=7)

    ws.cell(row=1, column=1, value=f"{project_title} — generated {today.isoformat()}").font = Font(bold=True, size=13)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill, cell.font, cell.border = head_fill, head_font, border
    for idx, week in enumerate(weeks):
        cell = ws.cell(row=2, column=n_static + 1 + idx, value=week.strftime("%d.%m"))
        cell.fill, cell.border = head_fill, border
        cell.font = Font(bold=True, color="FFFFFF", size=8)
        cell.alignment = Alignment(text_rotation=90, horizontal="center")
        ws.column_dimensions[get_column_letter(n_static + 1 + idx)].width = 3.2

    for r, item in enumerate(rows, start=3):
        values = [
            f"#{item['number']}" if item["number"] else "",
            item["title"], item["area"], item["priority"], item["size"],
            item["status"],
            item["start"].isoformat() if item["start"] else "",
            item["finish"].isoformat() if item["finish"] else "",
            f"{item['pct']}%",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = border
        if item["url"]:
            ws.cell(row=r, column=1).hyperlink = item["url"]
            ws.cell(row=r, column=1).font = Font(color="1565C0", underline="single")

        # Gantt bar
        if item["start"]:
            finish = item["finish"] or max(item["start"], today)
            fill = PatternFill("solid", fgColor=STATUS_COLOR.get(item["status"], "90A4AE"))
            for idx, week in enumerate(weeks):
                if item["start"] <= week + dt.timedelta(days=6) and finish >= week:
                    ws.cell(row=r, column=n_static + 1 + idx).fill = fill
            # % label right after the bar
            last = max(i for i, w in enumerate(weeks) if item["start"] <= w + dt.timedelta(days=6) and finish >= w)
            label = ws.cell(row=r, column=n_static + 2 + last, value=f"{item['pct']}%")
            label.font = Font(size=8, bold=True)

    widths = [7, 62, 12, 9, 6, 12, 11, 11, 11]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=3, column=n_static + 1)

    wb.save(path)


def main() -> None:
    token = os.environ.get("GITHUB_TOKEN") or os.environ.get("GH_TOKEN")
    if not token:
        sys.exit("GITHUB_TOKEN is required (token with read:project scope)")
    owner = os.environ.get("PROJECT_OWNER", "YurMil")
    number = int(os.environ.get("PROJECT_NUMBER", "2"))
    output = os.environ.get("OUTPUT_FILE", "project-report.xlsx")

    title, nodes = fetch_items(token, owner, number)
    rows = [item for item in (parse_item(n) for n in nodes) if item]
    order = {"To triage": 0, "In progress": 1, "In review": 2, "Ready": 3, "Backlog": 4, "Done": 5}
    rows.sort(key=lambda x: (order.get(x["status"], 9), x["number"] or 0))

    build_workbook(title, rows, output)
    print(f"Wrote {output}: {len(rows)} tasks from project '{title}'")


if __name__ == "__main__":
    main()
